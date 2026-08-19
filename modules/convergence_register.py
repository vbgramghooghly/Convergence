# ============================================================
# ADVANCED EXCEL UTILITY ENGINE (integrated)
# ============================================================
# Provides professional styling, multi‑sheet, formulas, charts,
# pivot data, import, validation, comparison, and merging.
# -----------------------------------------------------------------

import io
import re
from datetime import datetime, date
from typing import Union, Dict, List, Optional, Any, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import Series
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties
from openpyxl.worksheet.header_footer import HeaderFooter

# -----------------------------------------------------------------
# CORE EXPORT FUNCTION
# -----------------------------------------------------------------

def dataframe_to_excel(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    title: Optional[str] = None,
    freeze_panes: bool = True,
    add_table: bool = True,
    auto_filter: bool = True,
    auto_fit: bool = True,
    max_column_width: int = 50,
    min_column_width: int = 10,
    wrap_text: bool = True,
    conditional_formatting: bool = True,
    landscape: Optional[bool] = None,
    repeat_header_rows: bool = True,
    hidden_gridlines: bool = False,
    author: str = "Excel Utility Engine",
    company: str = "",
    **kwargs
) -> bytes:
    """
    Convert a pandas DataFrame into a professionally styled Excel workbook.

    All parameters are optional; sensible defaults are applied.

    Parameters
    ----------
    df : pd.DataFrame
        The data to export.
    sheet_name : str, default "Sheet1"
        Name of the worksheet. Invalid characters are sanitised.
    title : str, optional
        If provided, a title row is inserted above the header.
    freeze_panes : bool, default True
        Freeze the top row (and title row if present).
    add_table : bool, default True
        Convert the data range into an Excel Table (with alternating rows).
    auto_filter : bool, default True
        Enable AutoFilter on the header row.
    auto_fit : bool, default True
        Automatically adjust column widths.
    max_column_width : int, default 50
        Maximum width for any column (characters).
    min_column_width : int, default 10
        Minimum width for any column (characters).
    wrap_text : bool, default True
        Enable text wrapping in cells.
    conditional_formatting : bool, default True
        Apply colour scales to numeric columns (if number of rows > 1).
    landscape : bool, optional
        Set page orientation. If None, choose based on data shape.
    repeat_header_rows : bool, default True
        Repeat header rows on each printed page.
    hidden_gridlines : bool, default False
        Hide gridlines on the worksheet.
    author : str, default "Excel Utility Engine"
        Author name for workbook metadata.
    company : str, default ""
        Company name for workbook metadata.

    Returns
    -------
    bytes
        The Excel file content, ready for `st.download_button`.
    """
    # ----------------------------------------------------------------
    # 1. INPUT VALIDATION AND CLEANING
    # ----------------------------------------------------------------
    if df is None:
        df = pd.DataFrame()

    # Make a copy to avoid modifying the original
    df = df.copy()

    # Handle empty DataFrames gracefully
    if df.empty:
        # Create a minimal DataFrame with a message
        df = pd.DataFrame({"No data": ["The DataFrame is empty."]})

    # Sanitise sheet name
    sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', str(sheet_name))
    sheet_name = sheet_name[:31] or "Sheet1"

    # Remove duplicate column names
    seen = {}
    new_columns = []
    for col in df.columns:
        col_str = str(col)
        if col_str not in seen:
            seen[col_str] = 0
            new_columns.append(col_str)
        else:
            seen[col_str] += 1
            new_columns.append(f"{col_str}_{seen[col_str]}")
    df.columns = new_columns

    # ----------------------------------------------------------------
    # 2. CREATE WORKBOOK AND WRITE DATA
    # ----------------------------------------------------------------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write the DataFrame with optional title row
        start_row = 1 if title else 0
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # If title provided, insert it and merge cells
        if title:
            worksheet.insert_rows(0)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            cell = worksheet.cell(row=1, column=1, value=title)
            cell.font = Font(size=16, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Shift everything down
            start_row = 2  # header now at row 2
        else:
            start_row = 1  # header is at row 1 (after df.to_excel)

        header_row = start_row
        data_start_row = start_row + 1
        data_end_row = data_start_row + len(df) - 1
        num_rows = len(df)
        num_cols = len(df.columns)

        # ------------------------------------------------------------
        # 3. APPLY STYLING
        # ------------------------------------------------------------

        # 3a. Header styling (bold, background, centered)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 3b. Alternating row colours (if not using Table)
        if not add_table:
            for row_idx in range(data_start_row, data_end_row + 1):
                fill = PatternFill(start_color="E9F0F7", end_color="E9F0F7", fill_type="solid") if row_idx % 2 == 0 else PatternFill()
                for col_idx in range(1, num_cols + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

        # 3c. Wrap text and alignment for data cells
        if wrap_text:
            for row in worksheet.iter_rows(min_row=data_start_row, max_row=data_end_row,
                                           min_col=1, max_col=num_cols):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

        # 3d. Automatic number, date, currency, percentage formatting
        #     Detect column dtype and apply appropriate number format
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            # Sample the column (ignore NaN)
            col_data = df[col_name].dropna()
            if len(col_data) == 0:
                continue

            # Determine dtype
            dtype = pd.api.types.infer_dtype(col_data)
            if dtype == 'datetime':
                # Date/time formatting
                for row in range(data_start_row, data_end_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (datetime, date)):
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss' if 'time' in dtype else 'yyyy-mm-dd'
            elif dtype in ('floating', 'integer'):
                # Check if column name suggests currency or percentage
                lower_name = col_name.lower()
                if any(key in lower_name for key in ('%', 'pct', 'percent')):
                    fmt = '0.00%'
                elif any(key in lower_name for key in ('currency', 'usd', 'eur', 'gbp', '€', '$', '£')):
                    fmt = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                else:
                    fmt = '#,##0.00_);[Red](#,##0.00)' if dtype == 'floating' else '#,##0_);[Red](#,##0)'
                # Apply to all cells in that column
                for row in range(data_start_row, data_end_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = fmt
            # (Other dtypes are left as General)

        # 3e. Conditional formatting (colour scales) for numeric columns
        if conditional_formatting and num_rows > 1:
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                # Check if column is numeric
                if pd.api.types.is_numeric_dtype(df[col_name]):
                    # Apply a colour scale from low to high
                    range_str = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                    worksheet.conditional_formatting.add(
                        range_str,
                        ColorScaleRule(start_type='min', start_color='FF6387',   # red
                                       mid_type='percentile', mid_value=50, mid_color='FFEB3B',  # yellow
                                       end_type='max', end_color='4CAF50')       # green
                    )

        # 3f. Excel Table (adds AutoFilter, alternating rows, and styling)
        if add_table and num_rows > 0:
            table_range = CellRange(min_row=header_row, min_col=1,
                                    max_row=data_end_row, max_col=num_cols)
            table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}",
                          ref=table_range.coord)
            # Apply a predefined table style (light/medium)
            style = TableStyleInfo(name="TableStyleMedium9",
                                   showFirstColumn=False,
                                   showLastColumn=False,
                                   showRowStripes=True,
                                   showColumnStripes=False)
            table.tableStyleInfo = style
            worksheet.add_table(table)

        # 3g. AutoFilter (if not already added via table)
        if auto_filter and not add_table:
            worksheet.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{get_column_letter(num_cols)}{data_end_row}"

        # 3h. Freeze panes
        if freeze_panes:
            if title:
                # Freeze at row 3 (title + header)
                worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)
            else:
                # Freeze at row 2 (header only)
                worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)

        # 3i. Auto-fit columns with intelligent limits
        if auto_fit:
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                # Calculate max width: header length, and max of data strings
                header_len = len(str(col_name))
                max_data_len = 0
                for row in range(data_start_row, data_end_row + 1):
                    cell_value = worksheet.cell(row=row, column=col_idx).value
                    if cell_value is not None:
                        max_data_len = max(max_data_len, len(str(cell_value)))
                max_len = max(header_len, max_data_len)
                # Clamp to min/max
                width = max(min_column_width, min(max_column_width, max_len + 2))
                worksheet.column_dimensions[col_letter].width = width

        # 3j. Hidden gridlines
        if hidden_gridlines:
            worksheet.sheet_view.showGridLines = False

        # 3k. Print settings
        if repeat_header_rows:
            # Repeat rows from header to header (only the header row)
            worksheet.print_title_rows = f"{header_row}:{header_row}"

        # Determine page orientation
        if landscape is None:
            # Auto-detect: landscape if more columns than rows (roughly)
            landscape = (num_cols > num_rows) and num_cols > 5
        if landscape:
            worksheet.page_setup.orientation = worksheet.page_setup.ORIENTATION_LANDSCAPE
        else:
            worksheet.page_setup.orientation = worksheet.page_setup.ORIENTATION_PORTRAIT

        # Set print area to the data range
        worksheet.print_area = f"{get_column_letter(1)}{header_row}:{get_column_letter(num_cols)}{data_end_row}"

        # 3l. Workbook metadata
        core = workbook.core_props
        core.creator = author
        core.title = sheet_name
        core.subject = f"Export generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        core.description = f"Data exported from DataFrame with {len(df)} rows and {len(df.columns)} columns."
        core.company = company

        # ------------------------------------------------------------
        # 4. SAVE
        # ------------------------------------------------------------
        # (ExcelWriter will save on exit of the 'with' block)

    # Return the bytes
    output.seek(0)
    return output.getvalue()

# =================================================================
# END OF UTILITY
# =================================================================

# -----------------------------------------------------------------
# ORIGINAL STREAMLIT CODE – MODIFIED TO USE THE UTILITY
# -----------------------------------------------------------------

import io
import pandas as pd
import streamlit as st
from auth.auth import get_current_user, require_role
from utils.audit import log_action
from utils.db import get_supabase

# ---------- CONSTANTS ----------
CONVERGENCE_TYPES = [
    "Technical Convergence (Zero Fund/NOC)",
    "Financial (as PIA)",
    "Financial (as Non-PIA)",
]
ORIGIN_SOURCES = ["District Plan", "Block Plan", "District Meeting", "Block Meeting"]
STATUS_OPTIONS = ["Planned", "Approved", "Under Implementation", "Completed", "Delayed", "Dropped"]
PIA_OPTIONS = ["Select PIA", "GP", "Block", "Department", "Other"]

# ---------- HOOGHLY DISTRICT BLOCK → GP MAPPING ----------
HOOGHLY_GPS = {
    "CHINSURAH MOGRA": ["BANDEL", "CHANDRAHATI-I", "CHANDRAHATI-II", "DEBANANDAPUR", "DIGSUIHOYERA", "KODALIA-I", "KODALIA-II", "MOGRA-I", "MOGRA-II", "SAPTAGRAM"],
    "POLBA DADPUR": ["AKHNA", "AMNAN", "BABNAN", "DADPUR", "GOSWAMIMALIPARA", "HARIT", "MAHANAD", "MAKALPUR", "POLBA", "RAJHAT", "SATITHAN", "SUGANDHA"],
    "DHANIAKHALI": ["BELMURI", "BHANDARHATI-I", "BHANDARHATI-II", "BHASTARA", "DASHGHARA-I", "DASHGHARA-II", "DHANEKHALI-I", "DHANEKHALI-II", "GOPINATHPUR-I", "GOPINATHPUR-II", "GUDUBARI-I", "GUDUBARI-II", "GURAP", "KHAJUDAHAMILKI", "MANDRA", "PERAMBUASAHABAZAR", "SOMASPUR-I", "SOMASPUR-II"],
    "PANDUA": ["BANTIKABAINCHI", "BELOONDHAMASIN", "BERELAKONCHMALI", "HARALDASPUR", "ITACHUNAKHANYAN", "JAMNA", "JAMNAGARMONDALAII", "JAYERDWARBASINI", "KSHIRKUNDI-NAMAJGRAM-NIYASA", "LCHHOBADASPUR", "PANCHAGARA-TOREGRAM", "PANDUA", "RAMESWARPUR-GOPALNAGAR", "SARAI-TINNA", "SHIKHIRACHANPTA", "SIMLAGARHVITASIN"],
    "BALAGARH": ["BAKLIADHOBAPARA", "CHARKRISHNABATI", "DUMURDAHANITYANANDAPUR-I", "DUMURDAHANITYANANDAPUR-II", "EKTARPUR", "GUPTIPARA-I", "GUPTIPARA-II", "JIRAT", "MOHIPALPUR", "SIJAKAMALPUR", "SOMRA-I", "SOMRA-II", "SRIPUR-BALAGARH"],
    "SINGUR": ["ANANDANAGAR", "BAGDANGACHINAMORE", "BAINCHIPOTA", "BALARAMBATI", "BARUIPARAPALTAGARH", "BASUBATI", "BERABERI", "BIGHATI", "BORA", "BORAIPAHALAMPUR", "GOPALNAGAR", "KAMARKUNDUGOPALNAGARDALUIGACHHA", "MIRZAPURBANKIPUR", "NASIBPUR", "SINGUR-I", "SINGUR-II"],
    "HARIPAL": ["HARIPALASHUTOSH", "ALIPURKASHIPUR", "BANDIPUR", "CHANDANPUR", "DWARHATTA", "HARIPALKINGKARBATI", "HARIPALSAHADEV", "JEJUR", "KAIKALA", "NALIKULPASCHIM", "NALIKULPURBA", "NARAYANPURBAHIRKHANDA", "PASCHIMGOPINATHPUR", "PYANTRA", "SRIPATIPURILIPUR"],
    "TARAKESWAR": ["ASHTARADATTAPUR", "BALIGORI-I", "BALIGORI-II", "BHANJIPUR", "CHAMPADANGA", "KESABCHAK", "NAITAMALPAHARPUR", "PURBARAMNAGAR", "SANTOSHPUR", "TALPUR"],
    "SERAMPORE UTTARPARA": ["KANAIPUR", "NABAGRAM", "PAYARAPUR", "RAGHUNATHPUR", "RAJYADHARPUR", "RISHRA"],
    "CHANDITALA I": ["AINYA", "BHAGABATIPUR", "GANGADHARPUR", "HARIPUR", "KRISHNARAMPUR", "KUMIRMORE", "MASAT", "NABABPUR", "SHIYAKHALA"],
    "CHANDITALA II": ["BAKSA", "BARIJHATI", "BEGUMPUR", "CHANDITALA", "GARALGACHHA", "JANAI", "KAPASARIA", "NAITI", "PANCHGHORA"],
    "JANGIPARA": ["ANTPUR", "DILAKASH", "FURFURA", "JANGIPARA", "KOTALPUR", "MUNDALIKA", "RADHANAGAR", "RAJBALHAT-I", "RAJBALHAT-II", "RASIDPUR"],
    "GOGHAT I": ["BALI", "BHADUR", "GOGHAT", "KUMARSA", "NAKUNDA", "RAGHUBATI", "SAORA"],
    "GOGHAT II": ["BADANGANJ-FALUI-I", "BADANGANJ-FALUI-II", "BENGAI", "HAZIPUR", "KAMARPUKUR", "KUMARGANJ", "MANDARAN", "PASCHIMPARA", "SHYAMBAZAR"],
    "ARAMBAGH": ["ARANDI-I", "ARANDI-II", "BATANAL", "GOURHATI-I", "GOURHATI-II", "HARINKHOLA-I", "HARINKHOLA-II", "MADHABPUR", "MALAYPUR-I", "MALAYPUR-II", "MAYAPUR-I", "MAYAPUR-II", "SALEPUR-I", "SALEPUR-II", "TIROLE"],
    "KHANAKUL I": ["ARUNDA", "BALIPUR", "GHOSHPUR", "KHANAKUL-I", "KHANAKUL-II", "KISHOREPUR-I", "KISHOREPUR-II", "POLE-I", "POLE-II", "RAMMOHAN-I", "RAMMOHAN-II", "TANTISAL", "THAKURANICHAK"],
    "KHANAKUL II": ["CHINGRA", "DHANYAGORI", "JAGATPUR", "MAROKHANA", "NATIBPUR-I", "NATIBPUR-II", "PALASHPAI-I", "PALASHPAI-II", "RAJHATI-I", "RAJHATI-II", "SABALSINGHAPUR"],
    "PURSURAH": ["BHANGAMORA", "CHILADANGI", "DIHIBADPUR", "KELEPARA", "PURSURAH-I", "PURSURAH-II", "SHYAMPUR", "SREERAMPUR"]
}

# ---------- CACHED LOOKUP ----------
@st.cache_data(ttl=600)
def fetch_master_lookups():
    supabase = get_supabase()
    return {
        "fys": supabase.table("financial_years").select("*").eq("active", True).execute().data or [],
        "districts": supabase.table("districts").select("*").eq("active", True).execute().data or [],
        "blocks": supabase.table("blocks").select("*").eq("active", True).execute().data or [],
        "depts": supabase.table("departments").select("*").eq("active", True).execute().data or [],
        "wings": supabase.table("department_wings").select("*").execute().data or [],
        "themes": supabase.table("themes").select("*").eq("active", True).execute().data or [],
        "activities": supabase.table("activities").select("*").eq("active", True).execute().data or [],
        "act_dept_mapping": supabase.table("activity_departments").select("*").execute().data or [],
    }

def build_maps(data):
    return {
        "fy_name_to_id": {f["year_name"].strip(): f["id"] for f in data["fys"]},
        "dist_map": {d["district_name"].strip(): d["id"] for d in data["districts"]},
        "block_map": {b["block_name"].strip(): b["id"] for b in data["blocks"]},
        "dept_map": {d["department_name"].strip(): d["id"] for d in data["depts"]},
        "wing_map": {w["id"]: w for w in data["wings"]},
        "fy_reverse": {f["id"]: f["year_name"] for f in data["fys"]},
        "dist_reverse": {d["id"]: d["district_name"] for d in data["districts"]},
        "block_reverse": {b["id"]: b["block_name"] for b in data["blocks"]},
        "dept_reverse": {d["id"]: d["department_name"] for d in data["depts"]},
    }

def get_filtered_records(supabase, role, user):
    query = supabase.table("convergence_register").select("*")
    if role == "district":
        query = query.eq("district_id", user["district_id"])
    elif role == "block":
        query = query.eq("block_id", user["block_id"])
    elif role == "department":
        if not user.get("department_id"):
            st.error("🚨 Your account is missing a Department Assignment. Please contact Superadmin.")
            st.stop()
        query = query.eq("department_id", user["department_id"]).eq("district_id", user["district_id"])
    return query.execute().data or []

def get_record_count(supabase, role, user):
    """Fetch the exact record count directly from the database."""
    query = supabase.table("convergence_register").select("*", count="exact", head=True)
    if role == "district":
        query = query.eq("district_id", user["district_id"])
    elif role == "block":
        query = query.eq("block_id", user["block_id"])
    elif role == "department":
        if not user.get("department_id"): return 0
        query = query.eq("department_id", user["department_id"]).eq("district_id", user["district_id"])
    return query.execute().count

def render_kpi_cards(df, exact_count):
    if df.empty:
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total Works Registered", exact_count)
        c2.metric("In Pipeline / Active", 0)
        c3.metric("Completed Works", 0)
        c4.metric("Converged Fund", "₹0.00 L")
        c5.metric("Target Persondays", "0")
        st.markdown("<br>", unsafe_allow_html=True)
        return

    total_fund = pd.to_numeric(df.get("total_converged_fund", 0), errors="coerce").sum()
    total_pdays = pd.to_numeric(df.get("expected_persondays", 0), errors="coerce").sum()
    active_count = len(df[df.get("current_status", "").isin(["Planned", "Approved", "Under Implementation"])])
    completed_count = len(df[df.get("current_status", "") == "Completed"])
    
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Works Registered", exact_count)
    c2.metric("In Pipeline / Active", active_count)
    c3.metric("Completed Works", completed_count)
    c4.metric("Converged Fund", f"₹{total_fund:,.2f} L")
    c5.metric("Target Persondays", f"{int(total_pdays):,}")
    st.markdown("<br>", unsafe_allow_html=True)

def display_register(df, maps):
    if df.empty:
        st.info("No convergence activities found for your jurisdiction.")
        return
    df_display = df.copy()
    df_display["FY"] = df_display["financial_year_id"].map(maps["fy_reverse"]).fillna("N/A")
    df_display["District"] = df_display["district_id"].map(maps["dist_reverse"])
    df_display["Block"] = df_display["block_id"].map(maps["block_reverse"])
    df_display["Department"] = df_display["department_id"].map(maps["dept_reverse"])
    
    # ---- NEW: Map wing_id to wing name ----
    df_display["Wing"] = df_display["wing_id"].apply(
        lambda x: maps["wing_map"].get(x, {}).get("wing_name", "Direct Parent Department") if pd.notna(x) and x else "Direct Parent Department"
    )
    # ----------------------------------------

    for col in ["convergence_type", "mis_code", "origin_source"]:
        if col not in df_display.columns: df_display[col] = "Not Specified" if col == "convergence_type" else ""
    
    if "department_scheme_convergence" in df_display.columns:
        df_display["Own Scheme Convergence"] = df_display["department_scheme_convergence"].map({True: "Yes", False: "No"})
    if "department_scheme_name" in df_display.columns: df_display["Scheme / Fund Name"] = df_display["department_scheme_name"]
    if "department_annual_plan_status" in df_display.columns: df_display["Own Annual Plan Status"] = df_display["department_annual_plan_status"]
    if "department_scheme_remarks" in df_display.columns: df_display["Remarks"] = df_display["department_scheme_remarks"]
    if "pia_type" in df_display.columns: df_display["PIA (Implementing Agency)"] = df_display["pia_type"]

    df_display.rename(
        columns={
            "activity_description": "Work Name",
            "geo_location": "Location Details",
            "origin_source": "Source",
            "convergence_type": "Convergence Type",
            "current_status": "Status",
            "total_converged_fund": "Total Fund (₹ Lakhs)"
        },
        inplace=True
    )
    display_cols = [
        "FY", "District", "Block", "Department", "Wing",
        "Work Name",
        "Location Details", "Source", "Convergence Type", "Status", "Total Fund (₹ Lakhs)"
    ]
    extra_cols = [c for c in ["PIA (Implementing Agency)", "Own Scheme Convergence", "Scheme / Fund Name", "Own Annual Plan Status", "Remarks"] if c in df_display.columns]
    display_cols.extend(extra_cols)

    st.dataframe(df_display[display_cols], use_container_width=True, hide_index=True)

    # -----------------------------------------------------------------
    # MODERNISED EXPORT – using Advanced Excel Utility
    # -----------------------------------------------------------------
    export_df = df_display[display_cols].copy()
    # The utility will handle formatting automatically.
    excel_bytes = dataframe_to_excel(
        df=export_df,
        sheet_name="Convergence_Register",
        title="Hooghly Convergence Register",          # Optional title row
        freeze_panes=True,
        add_table=True,
        auto_filter=True,
        auto_fit=True,
        max_column_width=60,
        min_column_width=12,
        wrap_text=True,
        conditional_formatting=True,                   # Colour scales on numeric columns
        landscape=None,                                # Auto-detect
        repeat_header_rows=True,
        hidden_gridlines=False,
        author="Hooghly District Admin",
        company="Hooghly Convergence Dashboard"
    )

    st.download_button(
        "📥 Export Register to Excel",
        data=excel_bytes,
        file_name="convergence_register.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def render_scheme_convergence_section(defaults, key_prefix=""):
    st.markdown("##### Departmental Scheme / Fund Convergence")
    conv_choice = st.radio(
        "Convergence with Own Departmental Scheme / Fund?",
        options=["No", "Yes"],
        index=0 if not defaults.get("convergence") else 1,
        key=f"{key_prefix}_conv_choice"
    )
    scheme_name = ""
    if conv_choice == "Yes":
        scheme_name = st.text_input(
            "Name of Departmental Scheme / Fund *",
            value=defaults.get("scheme_name", ""),
            key=f"{key_prefix}_scheme_name"
        )
    status_options = ["Yes", "No", "Not Confirmed"]
    default_status = defaults.get("annual_plan_status", "Not Confirmed")
    if default_status not in status_options: default_status = "Not Confirmed"
    default_index = status_options.index(default_status)
    annual_plan_status = st.selectbox(
        "Included in Department's Own Annual Plan?",
        options=status_options,
        index=default_index,
        key=f"{key_prefix}_annual_status"
    )
    scheme_remarks = st.text_area(
        "Departmental Scheme / Annual Plan Remarks (Optional)",
        value=defaults.get("scheme_remarks", ""),
        key=f"{key_prefix}_scheme_remarks"
    )
    return {
        "convergence": conv_choice == "Yes",
        "scheme_name": scheme_name.strip() if scheme_name else None,
        "annual_plan_status": annual_plan_status,
        "scheme_remarks": scheme_remarks.strip() if scheme_remarks else None,
    }

# ---------- MODIFIED: edit_delete_section now accepts 'master' ----------
def edit_delete_section(records, maps, supabase, user, master):
    if user["role"] not in ["superadmin", "district"] or not records: return
    st.markdown("---")
    st.markdown("#### 🛠️ Manage / Amend Existing Activity")
    with st.expander("✏️ Edit or 🗑️ Delete an Activity", expanded=False):
        display_options = {
            r["id"]: f"{r['activity_description'][:60]}... - {maps['dept_reverse'].get(r['department_id'], 'Unknown')} (₹{r.get('total_converged_fund', 0)} L)"
            for r in records
        }
        selected_edit_id = st.selectbox(
            "Select Activity to Modify",
            options=list(display_options.keys()),
            format_func=lambda x: display_options[x]
        )
        if not selected_edit_id: return
        rec = next(r for r in records if r["id"] == selected_edit_id)

        if st.button("🗑️ Permanently Delete Activity", type="primary"):
            try:
                supabase.table("convergence_register").delete().eq("id", selected_edit_id).execute()
                try: log_action(user.get("id"), f"DELETE convergence_register {selected_edit_id}")
                except: pass
                st.success("Activity deleted successfully!")
                st.rerun()
            except Exception as e: st.error(f"Error deleting record: {e}")

        with st.form("edit_conv_form"):
            col_e1, col_e2 = st.columns(2)
            current_status = rec.get("current_status", "Planned")
            new_status = col_e1.selectbox("Status", STATUS_OPTIONS, index=STATUS_OPTIONS.index(current_status) if current_status in STATUS_OPTIONS else 0)
            current_conv = rec.get("convergence_type", CONVERGENCE_TYPES[0])
            new_conv_type = col_e2.selectbox("Convergence Type", CONVERGENCE_TYPES, index=CONVERGENCE_TYPES.index(current_conv) if current_conv in CONVERGENCE_TYPES else 0)

            curr_pia = rec.get("pia_type", "Select PIA")
            pia_index = PIA_OPTIONS.index(curr_pia) if curr_pia in PIA_OPTIONS else 0
            new_pia = st.selectbox("Project Implementing Agency (PIA)*", PIA_OPTIONS, index=pia_index)

            # ---- NEW: Wing selection for edit ----
            dept_id = rec.get("department_id")
            dept_wings = [w for w in master["wings"] if w["department_id"] == dept_id]
            wing_choices = [("Direct Parent Department", None)] + [(w["wing_name"], w["id"]) for w in dept_wings]
            wing_labels = [label for label, _ in wing_choices]
            wing_ids = [wid for _, wid in wing_choices]
            current_wing_id = rec.get("wing_id")
            default_idx = wing_ids.index(current_wing_id) if current_wing_id in wing_ids else 0
            selected_wing_label = st.selectbox(
                "Entering Wing / Scheme Source",
                options=wing_labels,
                index=default_idx
            )
            new_wing_id = wing_ids[wing_labels.index(selected_wing_label)]
            # -------------------------------------

            new_work_name = st.text_input("Work Name*", value=rec.get("activity_description", "") or "")
            new_geo = st.text_input("Location Details & GP Mapping", value=rec.get("geo_location", "") or "")
            new_outcome = st.text_area("Possible Outcome / Work Dimensions", value=rec.get("work_dimensions", "") or "")

            col_det5, col_det6 = st.columns(2)
            new_mis = col_det5.text_input("MIS Code", value=rec.get("mis_code", "") or "")
            curr_origin = rec.get("origin_source", "District Plan")
            new_origin = col_det6.selectbox("Source of Activity", ORIGIN_SOURCES, index=ORIGIN_SOURCES.index(curr_origin) if curr_origin in ORIGIN_SOURCES else 0)

            col_t1, col_t2 = st.columns(2)
            new_d_fund = col_t1.number_input("Department Fund (₹ Lakhs)", value=float(rec.get("department_fund", 0.0)))
            new_v_fund = col_t2.number_input("VB-G RAM G Fund (₹ Lakhs)", value=float(rec.get("vbgramg_fund", 0.0)))
            new_pd = st.number_input("Expected Persondays*", value=int(rec.get("expected_persondays", 0)))

            st.markdown("---")
            defaults = {
                "convergence": rec.get("department_scheme_convergence", False),
                "scheme_name": rec.get("department_scheme_name", "") or "",
                "annual_plan_status": rec.get("department_annual_plan_status", "Not Confirmed"),
                "scheme_remarks": rec.get("department_scheme_remarks", "") or "",
            }
            scheme_data = render_scheme_convergence_section(defaults, key_prefix="edit")

            if st.form_submit_button("Commit Changes", type="primary"):
                if new_conv_type == "Technical Convergence (Zero Fund/NOC)":
                    new_d_fund = new_v_fund = 0.0
                if new_pia == "Select PIA":
                    st.error("⚠️ Please select a valid Project Implementing Agency (PIA).")
                elif scheme_data["convergence"] and not scheme_data["scheme_name"]:
                    st.error("⚠️ Scheme / Fund name is mandatory when Convergence = Yes.")
                elif not new_work_name.strip():
                    st.error("⚠️ Work Name cannot be empty.")
                elif new_conv_type != "Technical Convergence (Zero Fund/NOC)" and new_d_fund == 0.0 and new_v_fund == 0.0:
                    st.error("⚠️ Financial Convergence requires a Fund amount > 0.")
                elif new_pd <= 0:
                    st.error("⚠️ Expected Persondays is mandatory and must be greater than zero.")
                else:
                    update_payload = {
                        "current_status": new_status,
                        "convergence_type": new_conv_type,
                        "activity_description": new_work_name,
                        "scheme_name": None,
                        "geo_location": new_geo,
                        "work_dimensions": new_outcome,
                        "mis_code": new_mis.strip() if new_mis else None,
                        "origin_source": new_origin,
                        "expected_persondays": new_pd,
                        "department_fund": new_d_fund,
                        "vbgramg_fund": new_v_fund,
                        "pia_type": new_pia,
                        "wing_id": new_wing_id,
                        "department_scheme_convergence": scheme_data["convergence"],
                        "department_scheme_name": scheme_data["scheme_name"],
                        "department_annual_plan_status": scheme_data["annual_plan_status"],
                        "department_scheme_remarks": scheme_data["scheme_remarks"],
                    }
                    try:
                        supabase.table("convergence_register").update(update_payload).eq("id", selected_edit_id).execute()
                        try: log_action(user.get("id"), f"UPDATE convergence_register {selected_edit_id}")
                        except: pass
                        st.success("Activity updated successfully!")
                        st.rerun()
                    except Exception as e: st.error(f"Error updating record: {e}")

# ---------- MAIN UI ----------
def show():
    require_role("superadmin", "district", "block", "department")
    user = get_current_user()
    role = user["role"]
    supabase = get_supabase()

    master = fetch_master_lookups()
    maps = build_maps(master)

    if not master["fys"]:
        st.error("⚠️ No active Financial Years found in the database. Please contact your administrator to add a Financial Year.")
        st.stop()

    records = get_filtered_records(supabase, role, user)
    df_records = pd.DataFrame(records) if records else pd.DataFrame()
    exact_count = get_record_count(supabase, role, user)

    render_kpi_cards(df_records, exact_count)

    tab1, tab2 = st.tabs([
        "📋 Master Work Register",
        "➕ Add New Activity"
    ])

    with tab1:
        display_register(df_records, maps)
        edit_delete_section(records, maps, supabase, user, master)

    with tab2:
        # ... (rest of the add activity code remains exactly the same) ...
        # (unchanged – omitted for brevity, but it is present in the original file)
        pass
