# ============================================================
# ADVANCED EXCEL UTILITY ENGINE
# ============================================================
# A production‑ready, enterprise‑grade Excel toolkit for
# Streamlit and data applications. Provides consistent styling,
# multi‑sheet exports, summaries, formulas, charts, pivot data,
# import/validation, comparison, and merging.
# ============================================================

import io
import re
from datetime import datetime, date
from typing import Union, Dict, List, Optional, Any, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import Series
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties
from openpyxl.worksheet.header_footer import HeaderFooter


# -----------------------------------------------------------------
# CORE EXPORT FUNCTION – everything else builds on this
# -----------------------------------------------------------------

def dataframe_to_excel(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    title: Optional[str] = None,
    freeze_panes: bool = True,
    add_table: bool = True,
    auto_filter: bool = True,
    auto_fit: bool = True,
    max_column_width: int = 50,
    min_column_width: int = 10,
    wrap_text: bool = True,
    conditional_formatting: bool = True,
    landscape: Optional[bool] = None,
    repeat_header_rows: bool = True,
    hidden_gridlines: bool = False,
    author: str = "Excel Utility Engine",
    company: str = "",
    **kwargs
) -> bytes:
    """
    Convert a pandas DataFrame into a professionally styled Excel workbook.

    All parameters are optional; sensible defaults are applied.

    Parameters
    ----------
    df : pd.DataFrame
        The data to export.
    sheet_name : str, default "Sheet1"
        Name of the worksheet. Invalid characters are sanitised.
    title : str, optional
        If provided, a title row is inserted above the header.
    freeze_panes : bool, default True
        Freeze the top row (and title row if present).
    add_table : bool, default True
        Convert the data range into an Excel Table (with alternating rows).
    auto_filter : bool, default True
        Enable AutoFilter on the header row.
    auto_fit : bool, default True
        Automatically adjust column widths.
    max_column_width : int, default 50
        Maximum width for any column (characters).
    min_column_width : int, default 10
        Minimum width for any column (characters).
    wrap_text : bool, default True
        Enable text wrapping in cells.
    conditional_formatting : bool, default True
        Apply colour scales to numeric columns (if number of rows > 1).
    landscape : bool, optional
        Set page orientation. If None, choose based on data shape.
    repeat_header_rows : bool, default True
        Repeat header rows on each printed page.
    hidden_gridlines : bool, default False
        Hide gridlines on the worksheet.
    author : str, default "Excel Utility Engine"
        Author name for workbook metadata.
    company : str, default ""
        Company name for workbook metadata.

    Returns
    -------
    bytes
        The Excel file content, ready for `st.download_button`.
    """
    # ----------------------------------------------------------------
    # 1. INPUT VALIDATION AND CLEANING
    # ----------------------------------------------------------------
    if df is None:
        df = pd.DataFrame()

    # Make a copy to avoid modifying the original
    df = df.copy()

    # Handle empty DataFrames gracefully
    if df.empty:
        # Create a minimal DataFrame with a message
        df = pd.DataFrame({"No data": ["The DataFrame is empty."]})

    # Sanitise sheet name
    sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', str(sheet_name))
    sheet_name = sheet_name[:31] or "Sheet1"

    # Remove duplicate column names
    seen = {}
    new_columns = []
    for col in df.columns:
        col_str = str(col)
        if col_str not in seen:
            seen[col_str] = 0
            new_columns.append(col_str)
        else:
            seen[col_str] += 1
            new_columns.append(f"{col_str}_{seen[col_str]}")
    df.columns = new_columns

    # ----------------------------------------------------------------
    # 2. CREATE WORKBOOK AND WRITE DATA
    # ----------------------------------------------------------------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write the DataFrame with optional title row
        start_row = 1 if title else 0
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # If title provided, insert it and merge cells
        if title:
            worksheet.insert_rows(0)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            cell = worksheet.cell(row=1, column=1, value=title)
            cell.font = Font(size=16, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Shift everything down
            start_row = 2  # header now at row 2
        else:
            start_row = 1  # header is at row 1 (after df.to_excel)

        header_row = start_row
        data_start_row = start_row + 1
        data_end_row = data_start_row + len(df) - 1
        num_rows = len(df)
        num_cols = len(df.columns)

        # ------------------------------------------------------------
        # 3. APPLY STYLING
        # ------------------------------------------------------------

        # 3a. Header styling (bold, background, centered)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 3b. Alternating row colours (if not using Table)
        if not add_table:
            for row_idx in range(data_start_row, data_end_row + 1):
                fill = PatternFill(start_color="E9F0F7", end_color="E9F0F7", fill_type="solid") if row_idx % 2 == 0 else PatternFill()
                for col_idx in range(1, num_cols + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

        # 3c. Wrap text and alignment for data cells
        if wrap_text:
            for row in worksheet.iter_rows(min_row=data_start_row, max_row=data_end_row,
                                           min_col=1, max_col=num_cols):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

        # 3d. Automatic number, date, currency, percentage formatting
        #     Detect column dtype and apply appropriate number format
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            # Sample the column (ignore NaN)
            col_data = df[col_name].dropna()
            if len(col_data) == 0:
                continue

            # Determine dtype
            dtype = pd.api.types.infer_dtype(col_data)
            if dtype == 'datetime':
                # Date/time formatting
                for row in range(data_start_row, data_end_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (datetime, date)):
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss' if 'time' in dtype else 'yyyy-mm-dd'
            elif dtype in ('floating', 'integer'):
                # Check if column name suggests currency or percentage
                lower_name = col_name.lower()
                if any(key in lower_name for key in ('%', 'pct', 'percent')):
                    fmt = '0.00%'
                elif any(key in lower_name for key in ('currency', 'usd', 'eur', 'gbp', '€', '$', '£')):
                    fmt = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                else:
                    fmt = '#,##0.00_);[Red](#,##0.00)' if dtype == 'floating' else '#,##0_);[Red](#,##0)'
                # Apply to all cells in that column
                for row in range(data_start_row, data_end_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = fmt
            # (Other dtypes are left as General)

        # 3e. Conditional formatting (colour scales) for numeric columns
        if conditional_formatting and num_rows > 1:
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                # Check if column is numeric
                if pd.api.types.is_numeric_dtype(df[col_name]):
                    # Apply a colour scale from low to high
                    range_str = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                    worksheet.conditional_formatting.add(
                        range_str,
                        ColorScaleRule(start_type='min', start_color='FF6387',   # red
                                       mid_type='percentile', mid_value=50, mid_color='FFEB3B',  # yellow
                                       end_type='max', end_color='4CAF50')       # green
                    )

        # 3f. Excel Table (adds AutoFilter, alternating rows, and styling)
        if add_table and num_rows > 0:
            table_range = CellRange(min_row=header_row, min_col=1,
                                    max_row=data_end_row, max_col=num_cols)
            table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}",
                          ref=table_range.coord)
            # Apply a predefined table style (light/medium)
            style = TableStyleInfo(name="TableStyleMedium9",
                                   showFirstColumn=False,
                                   showLastColumn=False,
                                   showRowStripes=True,
                                   showColumnStripes=False)
            table.tableStyleInfo = style
            worksheet.add_table(table)

        # 3g. AutoFilter (if not already added via table)
        if auto_filter and not add_table:
            worksheet.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{get_column_letter(num_cols)}{data_end_row}"

        # 3h. Freeze panes
        if freeze_panes:
            if title:
                # Freeze at row 3 (title + header)
                worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)
            else:
                # Freeze at row 2 (header only)
                worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)

        # 3i. Auto-fit columns with intelligent limits
        if auto_fit:
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                # Calculate max width: header length, and max of data strings
                header_len = len(str(col_name))
                max_data_len = 0
                for row in range(data_start_row, data_end_row + 1):
                    cell_value = worksheet.cell(row=row, column=col_idx).value
                    if cell_value is not None:
                        max_data_len = max(max_data_len, len(str(cell_value)))
                max_len = max(header_len, max_data_len)
                # Clamp to min/max
                width = max(min_column_width, min(max_column_width, max_len + 2))
                worksheet.column_dimensions[col_letter].width = width

        # 3j. Hidden gridlines
        if hidden_gridlines:
            worksheet.sheet_view.showGridLines = False

        # 3k. Print settings
        if repeat_header_rows:
            # Repeat rows from header to header (only the header row)
            worksheet.print_title_rows = f"{header_row}:{header_row}"

        # Determine page orientation – FIXED: using string literals
        if landscape is None:
            # Auto-detect: landscape if more columns than rows (roughly)
            landscape = (num_cols > num_rows) and num_cols > 5
        if landscape:
            worksheet.page_setup.orientation = 'landscape'
        else:
            worksheet.page_setup.orientation = 'portrait'

        # Set print area to the data range
        worksheet.print_area = f"{get_column_letter(1)}{header_row}:{get_column_letter(num_cols)}{data_end_row}"

        # 3l. Workbook metadata (workbook.core_props is always available)
        core = workbook.core_props
        core.creator = author
        core.title = sheet_name
        core.subject = f"Export generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        core.description = f"Data exported from DataFrame with {len(df)} rows and {len(df.columns)} columns."
        core.company = company

        # ------------------------------------------------------------
        # 4. SAVE
        # ------------------------------------------------------------
        # (ExcelWriter will save on exit of the 'with' block)

    # Return the bytes
    output.seek(0)
    return output.getvalue()


# =================================================================
# EXTENDED UTILITY FUNCTIONS
# =================================================================

def export_dataframe(df: pd.DataFrame, **kwargs) -> bytes:
    """Alias for `dataframe_to_excel`."""
    return dataframe_to_excel(df, **kwargs)


def export_multi_sheet_workbook(
    dataframes: Dict[str, pd.DataFrame],
    title_per_sheet: Optional[Dict[str, str]] = None,
    **default_style_kwargs
) -> bytes:
    """
    Export multiple DataFrames into a single workbook, each on its own sheet.

    Parameters
    ----------
    dataframes : dict
        Mapping of sheet name -> DataFrame.
    title_per_sheet : dict, optional
        Mapping of sheet name -> title (optional).
    **default_style_kwargs
        Styling parameters passed to `dataframe_to_excel` for each sheet.

    Returns
    -------
    bytes
        Excel file content.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dataframes.items():
            # Sanitise sheet name
            safe_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', str(sheet_name))[:31] or "Sheet"
            # Apply title if provided
            title = title_per_sheet.get(sheet_name) if title_per_sheet else None
            # Generate a temporary Excel file for this sheet and copy it
            sheet_bytes = dataframe_to_excel(df, sheet_name=safe_name, title=title, **default_style_kwargs)
            temp_wb = load_workbook(io.BytesIO(sheet_bytes))
            temp_sheet = temp_wb.active
            # Copy sheet to main workbook
            if safe_name in writer.book.sheetnames:
                # If sheet name exists, make it unique
                safe_name = safe_name + "_1"
            new_sheet = writer.book.create_sheet(title=safe_name)
            for row in temp_sheet.iter_rows(values_only=False):
                for cell in row:
                    new_sheet[cell.coordinate].value = cell.value
                    new_sheet[cell.coordinate]._style = cell._style  # copy style
            # Copy column dimensions
            for col in temp_sheet.column_dimensions:
                new_sheet.column_dimensions[col] = temp_sheet.column_dimensions[col]
            # Copy print settings, etc. (simplified)
        # If no sheets, create an empty sheet
        if not writer.book.sheetnames:
            writer.book.create_sheet("Empty")
    output.seek(0)
    return output.getvalue()


def export_with_summary(
    df: pd.DataFrame,
    summary_stats: Optional[Dict[str, Any]] = None,
    summary_sheet_name: str = "Summary",
    **kwargs
) -> bytes:
    """
    Export a DataFrame with an additional summary sheet containing statistics.

    Parameters
    ----------
    df : pd.DataFrame
        Main data.
    summary_stats : dict, optional
        Custom summary to display. If None, auto-generate (count, mean, etc.).
    summary_sheet_name : str, default "Summary"
        Name of the summary sheet.
    **kwargs
        Passed to `dataframe_to_excel` for the main sheet.

    Returns
    -------
    bytes
        Excel file content.
    """
    # Prepare summary DataFrame
    if summary_stats is None:
        # Auto summary
        numeric_cols = df.select_dtypes(include=['number']).columns
        summary_data = {}
        if len(numeric_cols) > 0:
            summary_data['Count'] = df[numeric_cols].count()
            summary_data['Mean'] = df[numeric_cols].mean()
            summary_data['Std'] = df[numeric_cols].std()
            summary_data['Min'] = df[numeric_cols].min()
            summary_data['Max'] = df[numeric_cols].max()
        # Add total rows? Might be better to handle separately.
        # We'll create a DataFrame with stats as rows, columns as metrics.
        summary_df = pd.DataFrame(summary_data).T.reset_index().rename(columns={'index': 'Metric'})
    else:
        summary_df = pd.DataFrame(summary_stats)

    # Export multi-sheet
    sheets = {kwargs.get('sheet_name', 'Data'): df, summary_sheet_name: summary_df}
    return export_multi_sheet_workbook(sheets, **kwargs)


def export_with_totals(
    df: pd.DataFrame,
    total_row_label: str = "Total",
    total_columns: Optional[List[str]] = None,
    **kwargs
) -> bytes:
    """
    Export a DataFrame with a total row appended (sum of numeric columns).

    The total row is added directly to the data, so the table includes it.

    Parameters
    ----------
    df : pd.DataFrame
        Main data.
    total_row_label : str, default "Total"
        Label in the first column of the total row.
    total_columns : list of str, optional
        Columns to sum. If None, all numeric columns are summed.
    **kwargs
        Passed to `dataframe_to_excel`.

    Returns
    -------
    bytes
        Excel file content.
    """
    df_totals = df.copy()
    numeric_cols = total_columns if total_columns else df.select_dtypes(include=['number']).columns.tolist()
    total_row = {col: '' for col in df.columns}
    total_row[df.columns[0]] = total_row_label
    for col in numeric_cols:
        total_row[col] = df[col].sum()
    df_totals = pd.concat([df_totals, pd.DataFrame([total_row])], ignore_index=True)
    return dataframe_to_excel(df_totals, **kwargs)


def export_with_formulas(
    df: pd.DataFrame,
    formula_columns: Dict[str, str],
    **kwargs
) -> bytes:
    """
    Export a DataFrame with additional columns containing Excel formulas.

    Parameters
    ----------
    df : pd.DataFrame
        Main data.
    formula_columns : dict
        Mapping of new column name -> Excel formula string.
        The formula will be applied to every row; use placeholders like {row}?.
        Actually, we'll write the formula directly into the cell.
        We'll append these as new columns at the end.
    **kwargs
        Passed to `dataframe_to_excel`.

    Returns
    -------
    bytes
        Excel file content.
    """
    # We'll need to write the DataFrame and then add formulas after.
    # Use a temporary writer? Better to use the core function but with a pre-processing step.
    # We'll create a new DataFrame that includes formulas as strings, but we need to
    # set the cell value as formula. For simplicity, we can generate the data and then
    # manually write formulas using openpyxl after export.
    # However, the core export uses pandas to write data. We can modify the writer.
    # Let's use the core function with a custom post-processing.
    # We'll write the DataFrame as usual, then open the workbook and add formula columns.
    # We'll create a wrapper.

    output = io.BytesIO()
    # First, write the base DataFrame without formulas
    base_bytes = dataframe_to_excel(df, **kwargs)
    wb = load_workbook(io.BytesIO(base_bytes))
    ws = wb.active

    # Add formula columns
    start_col = len(df.columns) + 1
    for col_name, formula in formula_columns.items():
        col_letter = get_column_letter(start_col)
        # Header
        ws.cell(row=1, column=start_col, value=col_name).font = Font(bold=True)
        # For each data row, write the formula
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            # Replace placeholders if needed (e.g., {A} with A2)
            # We'll assume the formula is already Excel-friendly with row references.
            # We'll replace {row} with the actual row number.
            formula_row = formula.replace("{row}", str(row_idx))
            ws.cell(row=row_idx, column=start_col, value=formula_row).number_format = 'General'
        start_col += 1

    # Save to bytes
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_with_charts(
    df: pd.DataFrame,
    chart_config: Dict[str, Any],
    **kwargs
) -> bytes:
    """
    Export a DataFrame with an embedded chart on a separate sheet.

    Parameters
    ----------
    df : pd.DataFrame
        Main data.
    chart_config : dict
        Configuration for the chart. Example:
        {
            'type': 'bar',          # or 'column', 'line'
            'x_axis': 'Category',   # column name for x-axis
            'y_axis': ['Sales', 'Profit'],  # columns for y-axis
            'title': 'Sales vs Profit',
            'sheet_name': 'Chart'
        }
    **kwargs
        Passed to `dataframe_to_excel` for the data sheet.

    Returns
    -------
    bytes
        Excel file content.
    """
    # Write data sheet using core function
    data_bytes = dataframe_to_excel(df, **kwargs)
    wb = load_workbook(io.BytesIO(data_bytes))
    ws_data = wb.active

    # Create chart sheet
    chart_sheet_name = chart_config.get('sheet_name', 'Chart')
    ws_chart = wb.create_sheet(title=chart_sheet_name[:31])

    # Determine chart type
    chart_type = chart_config.get('type', 'bar')
    if chart_type in ('bar', 'column'):
        chart = BarChart()
    else:
        chart = BarChart()  # fallback

    # Data references
    x_col = chart_config.get('x_axis')
    y_cols = chart_config.get('y_axis', [])
    if not x_col or not y_cols:
        # Fallback: use first column as x, rest as y
        x_col = df.columns[0]
        y_cols = df.columns[1:].tolist()

    # Find column indices
    col_map = {col: idx+1 for idx, col in enumerate(df.columns)}
    x_col_idx = col_map[x_col]
    y_col_indices = [col_map[y] for y in y_cols if y in col_map]

    # Add data to chart
    data_rows = len(df) + 1  # including header
    # x-values
    x_values = Reference(ws_data, min_col=x_col_idx, min_row=2, max_row=data_rows)
    # y-values
    for y_idx in y_col_indices:
        y_values = Reference(ws_data, min_col=y_idx, min_row=2, max_row=data_rows)
        series = Series(y_values, x_values, title_from_data=False)
        series.title = df.columns[y_idx-1]
        chart.append(series)

    # Set chart properties
    if chart_config.get('title'):
        chart.title = chart_config['title']
    chart.x_axis.title = x_col
    chart.y_axis.title = 'Values'
    chart.legend.position = 'b'

    # Add chart to the chart sheet
    ws_chart.add_chart(chart, 'A1')

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_with_pivot_data(
    df: pd.DataFrame,
    pivot_config: Dict[str, Any],
    **kwargs
) -> bytes:
    """
    Export a DataFrame with a separate sheet containing a pivot table
    (as static data, not a live pivot cache).

    Parameters
    ----------
    df : pd.DataFrame
        Main data.
    pivot_config : dict
        Configuration for pivot table:
        {
            'index': ['Region'],
            'columns': ['Year'],
            'values': ['Sales'],
            'aggfunc': 'sum',   # or 'mean', 'count', etc.
            'sheet_name': 'Pivot'
        }
    **kwargs
        Passed to `dataframe_to_excel` for the data sheet.

    Returns
    -------
    bytes
        Excel file content.
    """
    # Create pivot table using pandas
    pivot_df = df.pivot_table(
        index=pivot_config.get('index', []),
        columns=pivot_config.get('columns', []),
        values=pivot_config.get('values', []),
        aggfunc=pivot_config.get('aggfunc', 'sum')
    )
    # Reset index to make it a flat table
    pivot_df = pivot_df.reset_index()
    # Flatten columns
    pivot_df.columns = ['_'.join(map(str, col)).strip('_') for col in pivot_df.columns]

    # Now export multi-sheet with the pivot data
    sheets = {kwargs.get('sheet_name', 'Data'): df, pivot_config.get('sheet_name', 'Pivot'): pivot_df}
    return export_multi_sheet_workbook(sheets, **kwargs)


def export_template(
    template_headers: List[str],
    sheet_name: str = "Template",
    **kwargs
) -> bytes:
    """
    Generate an empty Excel template with headers and sample formatting.

    Parameters
    ----------
    template_headers : list
        Column names for the template.
    sheet_name : str, default "Template"
        Sheet name.
    **kwargs
        Passed to `dataframe_to_excel`.

    Returns
    -------
    bytes
        Excel file content.
    """
    df = pd.DataFrame(columns=template_headers)
    return dataframe_to_excel(df, sheet_name=sheet_name, **kwargs)


def import_excel(
    file_bytes: bytes,
    sheet_name: Optional[Union[str, int]] = None,
    **read_excel_kwargs
) -> Dict[str, pd.DataFrame]:
    """
    Import an Excel file and return a dict of DataFrames.

    Parameters
    ----------
    file_bytes : bytes
        The Excel file content.
    sheet_name : str or int, optional
        If provided, returns only that sheet as a single DataFrame in a dict.
        Otherwise, returns all sheets.
    **read_excel_kwargs
        Additional arguments passed to `pd.read_excel`.

    Returns
    -------
    dict
        Mapping of sheet name -> DataFrame.
    """
    if sheet_name is not None:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, **read_excel_kwargs)
        return {str(sheet_name): df}
    else:
        excel_data = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, **read_excel_kwargs)
        return {str(k): v for k, v in excel_data.items()}


def validate_excel(
    df: pd.DataFrame,
    required_columns: Optional[List[str]] = None,
    column_types: Optional[Dict[str, type]] = None,
    **kwargs
) -> Tuple[bool, List[str]]:
    """
    Validate a DataFrame against expected schema.

    Parameters
    ----------
    df : pd.DataFrame
        Data to validate.
    required_columns : list, optional
        Columns that must be present.
    column_types : dict, optional
        Mapping of column name -> expected dtype.
    **kwargs
        Additional validation rules (e.g., not_null, unique).

    Returns
    -------
    (is_valid, errors)
        is_valid : bool
        errors : list of strings describing issues.
    """
    errors = []
    if required_columns:
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            errors.append(f"Missing required columns: {missing}")
    if column_types:
        for col, expected_type in column_types.items():
            if col in df.columns:
                actual_dtype = df[col].dtype
                if not pd.api.types.is_dtype_equal(actual_dtype, expected_type):
                    errors.append(f"Column '{col}' has dtype {actual_dtype}, expected {expected_type}")
            else:
                errors.append(f"Column '{col}' not found for type validation")
    # Additional checks: not null
    if kwargs.get('not_null'):
        for col in kwargs['not_null']:
            if col in df.columns and df[col].isnull().any():
                errors.append(f"Column '{col}' contains null values")
    # Unique checks
    if kwargs.get('unique'):
        for col in kwargs['unique']:
            if col in df.columns and not df[col].is_unique:
                errors.append(f"Column '{col}' contains duplicate values")
    return len(errors) == 0, errors


def clean_excel(
    df: pd.DataFrame,
    strip_whitespace: bool = True,
    fillna: Optional[Dict[str, Any]] = None,
    dropna: bool = False,
    **kwargs
) -> pd.DataFrame:
    """
    Clean a DataFrame imported from Excel.

    Parameters
    ----------
    df : pd.DataFrame
        Raw data.
    strip_whitespace : bool, default True
        Strip leading/trailing whitespace from string columns.
    fillna : dict, optional
        Mapping of column -> value to fill NaN.
    dropna : bool, default False
        Drop rows with any NaN.
    **kwargs
        Additional cleaning steps.

    Returns
    -------
    pd.DataFrame
        Cleaned DataFrame.
    """
    df_clean = df.copy()
    if strip_whitespace:
        string_cols = df_clean.select_dtypes(include=['object']).columns
        df_clean[string_cols] = df_clean[string_cols].apply(lambda x: x.str.strip() if x.dtype == 'object' else x)
    if fillna:
        for col, value in fillna.items():
            if col in df_clean.columns:
                df_clean[col] = df_clean[col].fillna(value)
    if dropna:
        df_clean = df_clean.dropna()
    return df_clean


def compare_excel(
    file1_bytes: bytes,
    file2_bytes: bytes,
    sheet_name: Optional[Union[str, int]] = None,
    compare_kwargs: Optional[Dict[str, Any]] = None
) -> Dict[str, pd.DataFrame]:
    """
    Compare two Excel files and return differences.

    Parameters
    ----------
    file1_bytes, file2_bytes : bytes
        The two Excel files.
    sheet_name : str or int, optional
        Specific sheet to compare. If None, compare all sheets.
    compare_kwargs : dict, optional
        Passed to `pd.DataFrame.compare` (e.g., keep_equal, keep_shape).

    Returns
    -------
    dict
        Mapping of sheet name -> DataFrame with differences.
    """
    df1_dict = import_excel(file1_bytes, sheet_name=sheet_name)
    df2_dict = import_excel(file2_bytes, sheet_name=sheet_name)
    result = {}
    compare_kwargs = compare_kwargs or {}
    for sheet, df1 in df1_dict.items():
        if sheet in df2_dict:
            df2 = df2_dict[sheet]
            # Align indices and columns
            try:
                diff = df1.compare(df2, **compare_kwargs)
                if not diff.empty:
                    result[sheet] = diff
            except Exception as e:
                result[sheet] = pd.DataFrame({'Error': [str(e)]})
        else:
            result[sheet] = pd.DataFrame({'Info': ['Sheet not found in second file']})
    return result


def merge_excel(
    file_bytes_list: List[bytes],
    sheet_name: Optional[str] = None,
    axis: int = 0,
    **kwargs
) -> bytes:
    """
    Merge multiple Excel files (or sheets) into a single workbook.

    Parameters
    ----------
    file_bytes_list : list of bytes
        List of Excel file contents.
    sheet_name : str, optional
        If provided, only merge this specific sheet from each file.
    axis : int, default 0
        0 for vertical concatenation (rows), 1 for horizontal (columns).
    **kwargs
        Additional arguments for `pd.concat`.

    Returns
    -------
    bytes
        Merged Excel file content.
    """
    all_dfs = []
    for file_bytes in file_bytes_list:
        dfs = import_excel(file_bytes, sheet_name=sheet_name)
        # Take the first sheet if multiple and sheet_name not specified?
        if sheet_name is None:
            # Take the first sheet
            df = list(dfs.values())[0]
        else:
            df = dfs.get(sheet_name)
            if df is None:
                continue
        all_dfs.append(df)
    if not all_dfs:
        return dataframe_to_excel(pd.DataFrame())  # empty
    merged_df = pd.concat(all_dfs, axis=axis, **kwargs)
    return dataframe_to_excel(merged_df, sheet_name="Merged")


# =================================================================
# EXAMPLE USAGE IN STREAMLIT
# =================================================================
if __name__ == "__main__":
    # Demo - create a sample DataFrame and export
    sample_df = pd.DataFrame({
        'Product': ['Laptop', 'Mouse', 'Keyboard'],
        'Sales': [1200, 45, 78],
        'Profit (%)': [0.25, 0.10, 0.15],
        'Date': pd.date_range('2025-01-01', periods=3)
    })

    # Export to Excel bytes
    excel_bytes = dataframe_to_excel(
        sample_df,
        sheet_name="Products",
        title="Quarterly Sales Report",
        add_table=True,
        conditional_formatting=True
    )

    # In a Streamlit app, you would use:
    # st.download_button(
    #     label="📥 Download Excel",
    #     data=excel_bytes,
    #     file_name="report.xlsx",
    #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # )

    # To demonstrate multi-sheet:
    multi_bytes = export_multi_sheet_workbook(
        {
            'Sales': sample_df,
            'Summary': sample_df.describe().reset_index()
        },
        title_per_sheet={'Sales': 'Main Data'}
    )

    print("Excel generation successful. Bytes length:", len(excel_bytes))
    print("Multi-sheet bytes length:", len(multi_bytes))
